"""
Excel asset import parser
Generates downloadable templates and parses uploaded .xlsx files
into ParsedHost objects for the scan batch pipeline.
"""
import ipaddress
import logging
import re
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from app.utils.nmap_parser import ParsedHost, ParsedPort

logger = logging.getLogger(__name__)

# ── Column definitions ──────────────────────────────────────────

COLUMNS = [
    # (key, header_zh, header_en, required, comment_zh, comment_en)
    ("ip_address",       "IP地址",    "IP Address",     True,
     "必填，作为去重主键", "Required, used as dedup key"),
    ("mac_address",      "MAC地址",   "MAC Address",    False,
     "选填，格式 AA:BB:CC:DD:EE:FF", "Optional, format AA:BB:CC:DD:EE:FF"),
    ("hostname",         "主机名",    "Hostname",       False,
     "选填", "Optional"),
    ("os_info",          "操作系统",   "OS",             False,
     "选填，如 CentOS 7.9 / Windows Server 2019", "Optional, e.g. CentOS 7.9"),
    ("asset_type",       "资产类型",   "Asset Type",     False,
     "可选值: physical / virtual / network_device / other / cloud_server",
     "Values: physical / virtual / network_device / other / cloud_server"),
    ("network_zone",     "网络区域",   "Network Zone",   False,
     "可选值: dmz / intranet / office / management / other / aliyun / tencent / huawei / aws / azure / gcp / other_cloud",
     "Values: dmz / intranet / office / management / other / ..."),
    ("location",         "物理位置",   "Location",       False,
     "选填，如 机房A-3楼-C12", "Optional, e.g. DC-A-3F-C12"),
    ("owner",            "负责人",     "Owner",          False,
     "选填", "Optional"),
    ("business_system",  "业务系统",   "Business System", False,
     "选填", "Optional"),
    ("importance",       "重要性",     "Importance",     False,
     "可选值: core / important / normal", "Values: core / important / normal"),
    ("cpu",              "CPU",       "CPU",            False,
     "选填，如 8核", "Optional, e.g. 8 cores"),
    ("memory_gb",        "内存(GB)",   "Memory (GB)",    False,
     "选填，正整数", "Optional, positive integer"),
    ("disk_gb",          "磁盘(GB)",   "Disk (GB)",      False,
     "选填，正整数", "Optional, positive integer"),
    ("purchase_date",    "采购日期",   "Purchase Date",  False,
     "选填，格式 YYYY-MM-DD", "Optional, format YYYY-MM-DD"),
    ("warranty_expiry",  "保修到期",   "Warranty Expiry", False,
     "选填，格式 YYYY-MM-DD", "Optional, format YYYY-MM-DD"),
    ("port_number",      "端口号",     "Port Number",    False,
     "选填，1-65535", "Optional, 1-65535"),
    ("protocol",         "协议",       "Protocol",       False,
     "tcp 或 udp", "tcp or udp"),
    ("service_name",     "服务名称",   "Service Name",   False,
     "选填，如 nginx / mysql", "Optional, e.g. nginx"),
    ("remark",           "备注",       "Remark",         False,
     "自由文本，兜底字段", "Free text, fallback field"),
]

HEADER_KEYS = [c[0] for c in COLUMNS]

# ── Enum constraints ────────────────────────────────────────────

ASSET_TYPES = {"physical", "virtual", "network_device", "other", "cloud_server"}
NETWORK_ZONES = {
    "dmz", "intranet", "office", "management", "other",
    "aliyun", "tencent", "huawei", "aws", "azure", "gcp", "other_cloud",
}
IMPORTANCE_VALUES = {"core", "important", "normal"}
PROTOCOLS = {"tcp", "udp"}

# ── Data structures ─────────────────────────────────────────────

@dataclass
class ExcelRowError:
    """Validation error for a specific cell"""
    row_number: int      # 1-based Excel row (data starts at row 2)
    column: str          # column header name
    message: str


@dataclass
class ExcelParseResult:
    """Result of parsing an Excel file"""
    hosts: list[ParsedHost]
    errors: list[ExcelRowError]
    extra_fields: dict[str, dict] = field(default_factory=dict)  # ip -> supplementary fields

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0


# ── Template generation ─────────────────────────────────────────

def generate_template(language: str = "zh") -> bytes:
    """
    Generate a blank .xlsx template with headers and cell comments.
    language: 'zh' for Chinese headers, 'en' for English headers.
    Returns bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Import Template"

    # Header row
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_def in enumerate(COLUMNS, start=1):
        key, header_zh, header_en, required, comment_zh, comment_en = col_def

        # Choose header text based on language
        if language == "en":
            header_text = header_en
            comment_text = comment_en
            if required:
                header_text += " *"
        else:
            header_text = header_zh
            comment_text = comment_zh
            if required:
                header_text += " (必填)"

        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.comment = Comment(comment_text, "Z-CMDB", width=300, height=80)

        # Set column width
        ws.column_dimensions[cell.column_letter].width = max(14, len(header_text) * 2 + 4)

    # Example row
    example_data = {
        "ip_address": "192.168.1.100",
        "mac_address": "AA:BB:CC:DD:EE:FF",
        "hostname": "web-server-01",
        "os_info": "CentOS 7.9",
        "asset_type": "physical",
        "network_zone": "intranet",
        "location": "DC-A-3F-C12",
        "owner": "张三",
        "business_system": "企业官网",
        "importance": "important",
        "cpu": "8核",
        "memory_gb": "32",
        "disk_gb": "500",
        "purchase_date": "2024-01-15",
        "warranty_expiry": "2027-01-15",
        "port_number": "80",
        "protocol": "tcp",
        "service_name": "nginx",
        "remark": "主站 Web 服务器",
    }
    for col_idx, col_def in enumerate(COLUMNS, start=1):
        key = col_def[0]
        val = example_data.get(key, "")
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(color="808080", italic=True, size=10)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Parsing ─────────────────────────────────────────────────────

def _clean_str(val) -> str | None:
    """Convert cell value to stripped string, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _validate_ip(ip_str: str) -> bool:
    """Check if string is a valid IPv4 address."""
    try:
        ipaddress.IPv4Address(ip_str)
        return True
    except (ipaddress.AddressValueError, ValueError):
        return False


def _validate_mac(mac_str: str) -> bool:
    """Check MAC address format: XX:XX:XX:XX:XX:XX or XX-XX-XX-XX-XX-XX."""
    return bool(re.match(r'^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$', mac_str))


def parse_excel(content: bytes) -> ExcelParseResult:
    """
    Parse .xlsx content into ParsedHost objects.
    Validates each row, collects ALL errors without raising.
    Returns ExcelParseResult with hosts, errors, and extra_fields.
    """
    errors: list[ExcelRowError] = []
    hosts_map: dict[str, ParsedHost] = {}  # ip -> ParsedHost
    extra_fields: dict[str, dict] = {}     # ip -> supplementary fields

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        errors.append(ExcelRowError(row_number=0, column="-", message=f"无法打开 Excel 文件: {e}"))
        return ExcelParseResult(hosts=[], errors=errors)

    ws = wb.active
    if ws is None:
        errors.append(ExcelRowError(row_number=0, column="-", message="Excel 文件无工作表"))
        return ExcelParseResult(hosts=[], errors=errors)

    # Read header row to build column index mapping
    header_map: dict[str, int] = {}  # key -> column index (0-based)
    for col_idx, cell in enumerate(ws[1]):
        if cell.value is None:
            continue
        raw_header = str(cell.value).strip()
        # Strip common suffixes like "(必填)" or "*"
        clean_header = re.sub(r'\s*[\(（]必填[\)）]\s*$', '', raw_header).strip()
        clean_header = re.sub(r'\s*[\(（]Required[\)）]\s*$', '', clean_header).strip()
        clean_header = re.sub(r'\s*\*\s*$', '', clean_header).strip()

        # Match to known columns
        for col_def in COLUMNS:
            key, header_zh, header_en = col_def[0], col_def[1], col_def[2]
            if clean_header in (header_zh, header_en, key):
                if key not in header_map:  # first match wins
                    header_map[key] = col_idx
                break

    # Check that at least IP column is found
    if "ip_address" not in header_map:
        errors.append(ExcelRowError(
            row_number=1, column="IP地址",
            message="未找到「IP地址」列，请使用平台提供的标准模板"
        ))
        return ExcelParseResult(hosts=[], errors=errors)

    # Process data rows (row 2 onwards)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Get IP address
        ip_col = header_map.get("ip_address")
        if ip_col is None or ip_col >= len(row):
            continue
        ip_raw = _clean_str(row[ip_col])
        if ip_raw is None:
            continue  # skip empty rows

        # Validate IP
        if not _validate_ip(ip_raw):
            errors.append(ExcelRowError(row_number=row_idx, column="IP地址",
                                        message=f"无效的 IPv4 地址: {ip_raw}"))
            continue

        # Helper to get cell value by key
        def get(key: str) -> str | None:
            col = header_map.get(key)
            if col is None or col >= len(row):
                return None
            return _clean_str(row[col])

        # Validate MAC (if provided)
        mac_raw = get("mac_address")
        if mac_raw and not _validate_mac(mac_raw):
            errors.append(ExcelRowError(row_number=row_idx, column="MAC地址",
                                        message=f"无效的 MAC 地址格式: {mac_raw}"))
            mac_raw = None

        # Validate enum fields
        asset_type = get("asset_type")
        if asset_type and asset_type not in ASSET_TYPES:
            errors.append(ExcelRowError(row_number=row_idx, column="资产类型",
                                        message=f"无效的资产类型: {asset_type}，可选值: {', '.join(sorted(ASSET_TYPES))}"))
            asset_type = None

        network_zone = get("network_zone")
        if network_zone and network_zone not in NETWORK_ZONES:
            errors.append(ExcelRowError(row_number=row_idx, column="网络区域",
                                        message=f"无效的网络区域: {network_zone}"))
            network_zone = None

        importance = get("importance")
        if importance and importance not in IMPORTANCE_VALUES:
            errors.append(ExcelRowError(row_number=row_idx, column="重要性",
                                        message=f"无效的重要性: {importance}，可选值: core / important / normal"))
            importance = None

        # Validate numeric fields
        memory_gb = None
        mem_raw = get("memory_gb")
        if mem_raw:
            try:
                memory_gb = int(mem_raw)
                if memory_gb < 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(ExcelRowError(row_number=row_idx, column="内存(GB)",
                                            message=f"内存必须为非负整数: {mem_raw}"))
                memory_gb = None

        disk_gb = None
        disk_raw = get("disk_gb")
        if disk_raw:
            try:
                disk_gb = int(disk_raw)
                if disk_gb < 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(ExcelRowError(row_number=row_idx, column="磁盘(GB)",
                                            message=f"磁盘必须为非负整数: {disk_raw}"))
                disk_gb = None

        # Validate port
        port_number = None
        port_raw = get("port_number")
        protocol = None
        service_name = get("service_name")

        if port_raw:
            try:
                port_number = int(port_raw)
                if not (1 <= port_number <= 65535):
                    errors.append(ExcelRowError(row_number=row_idx, column="端口号",
                                                message=f"端口号必须在 1-65535 之间: {port_raw}"))
                    port_number = None
            except (ValueError, TypeError):
                errors.append(ExcelRowError(row_number=row_idx, column="端口号",
                                            message=f"端口号必须为整数: {port_raw}"))
                port_number = None

        if port_number:
            protocol = get("protocol") or "tcp"
            if protocol not in PROTOCOLS:
                errors.append(ExcelRowError(row_number=row_idx, column="协议",
                                            message=f"无效的协议: {protocol}，可选值: tcp / udp"))
                protocol = "tcp"  # default to tcp even if invalid

        # Get remark
        remark = get("remark")

        # Build or reuse ParsedHost
        if ip_raw not in hosts_map:
            hosts_map[ip_raw] = ParsedHost(
                ip_address=ip_raw,
                mac_address=mac_raw,
                hostname=get("hostname"),
                os_info=get("os_info"),
                ports=[],
            )
            # Store supplementary fields (first occurrence wins)
            extra_fields[ip_raw] = {
                "asset_type": asset_type,
                "network_zone": network_zone,
                "location": get("location"),
                "owner": get("owner"),
                "business_system": get("business_system"),
                "importance": importance,
                "cpu": get("cpu"),
                "memory_gb": memory_gb,
                "disk_gb": disk_gb,
                "purchase_date": get("purchase_date"),
                "warranty_expiry": get("warranty_expiry"),
                "remark": remark,
            }
        else:
            # Update mac/hostname/os if this row has them and the host doesn't
            host = hosts_map[ip_raw]
            if mac_raw and not host.mac_address:
                host.mac_address = mac_raw
            if get("hostname") and not host.hostname:
                host.hostname = get("hostname")
            if get("os_info") and not host.os_info:
                host.os_info = get("os_info")

        # Add port if specified
        if port_number:
            hosts_map[ip_raw].ports.append(ParsedPort(
                port_number=port_number,
                protocol=protocol,
                state="open",
                service_name=service_name,
                service_version=None,
            ))

    wb.close()

    hosts = list(hosts_map.values())
    logger.info("excel parsed: %d hosts, %d errors", len(hosts), len(errors))
    return ExcelParseResult(hosts=hosts, errors=errors, extra_fields=extra_fields)
